import time
import datetime
import openpyxl
from openpyxl.styles import Font

""" txt 파일을 엑셀 파일로 저장하기 """
import openpyxl
from openpyxl.styles import Font
from datetime import datetime
import re

# 엑셀 파일 생성
workbook = openpyxl.Workbook()
sheet = workbook.active

# 제목 행 작성
headers = ["Date", "Time", "Deviation direction", "Deviation percent"]
for col_num, header in enumerate(headers, 1):
    cell = sheet.cell(row=1, column=col_num, value=header)
    cell.font = Font(bold=True)

# 텍스트 파일에서 데이터 읽어오기
with open('10_13_test.txt', 'r') as text_file:
    lines = text_file.readlines()

# 정규 표현식 패턴 정의
pattern = r'(\w+) (\d+\.\d+)%'

# 각 줄의 데이터를 엑셀 파일에 추가
for row_num, line in enumerate(lines, 2):  # 2번째 행부터 시작
    data = line.strip().split(' - Deviation: ')
    if len(data) == 2:
        date, time = data[0].split()
        deviation_info = data[1]
        matches = re.findall(pattern, deviation_info)
        if matches:
            for match in matches:
                deviation_direction, deviation_percent = match
                sheet.cell(row=row_num, column=1, value=date)
                sheet.cell(row=row_num, column=2, value=time)
                sheet.cell(row=row_num, column=3, value=deviation_direction)
                sheet.cell(row=row_num, column=4, value=float(deviation_percent))

# 엑셀 파일 저장
workbook.save('10_13_test.xlsx')


""" ---------------------------------------------------------------------------------------------------------------- """


""" ---------------------------------------------------------------------------------------------------------------- """


# ----------------------------------------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------------------------

# # 파일에서 deviation 값만 읽어오는 함수
# def read_deviations_from_file(file_name):
#     deviations = []

#     # 파일을 읽기 모드로 열기
#     with open(file_name, 'r') as file:
#         for line in file:
#             if "Deviation:" in line:
#                 # "Deviation:" 문자열을 찾아서 해당 라인에서 deviation 값을 추출
#                 deviation = line.split("Deviation:")[1].strip()
#                 deviations.append(deviation)

#     return deviations

# # 1초마다 deviation을 txt 파일에 저장하는 함수
# def save_deviation_to_file(file_name):
#     start_time = time.time()
#     while True:
#         # 시간이 1초 경과하면 deviation을 계산하고 파일에 추가
#         current_time = time.time()
#         if current_time - start_time >= 1:
#             # deviation 계산 (예시로 간단히 0부터 9까지 순차적으로 저장)
#             deviation = str(current_time % 10)

#             # 현재 시간 얻기
#             current_time_str = time.strftime('%Y-%m-%d %H:%M:%S')

#             # deviation와 현재 시간을 함께 텍스트 파일에 추가 (append)
#             with open(file_name, 'a') as file:
#                 file.write(f"{current_time_str} - Deviation: {deviation}\n")  # 현재 시간과 deviation을 함께 저장

#             start_time = current_time  # 시작 시간 업데이트

# # -----------------------------------------------

# def read_deviations_from_file(file_name):
#     deviations = []

#     # 파일을 읽기 모드로 열기
#     with open(file_name, 'r') as file:
#         for line in file:
#             if "Deviation:" in line:
#                 # "Deviation:" 문자열을 찾아서 해당 라인에서 deviation 값을 추출
#                 deviation = line.split("Deviation:")[1].strip()
#                 deviations.append(deviation)

#     return deviations

# # deviation_project.txt 파일에서 deviation 값 읽어오기
# file_name = 'deviation_shadow2.txt'
# deviations = read_deviations_from_file(file_name)

# # 읽어온 deviation 값 출력
# for deviation in deviations:
#     print(deviation)
