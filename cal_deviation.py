import openpyxl
from openpyxl.styles import Font
from datetime import datetime

# 엑셀 파일 열기
workbook = openpyxl.load_workbook('10_13_test.xlsx')
sheet = workbook.active

# deviation 값을 모두 합산하여 deviation_sum에 저장
deviation_sum = 0
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=4, max_col=4):
    for cell in row:
        try:
            # 백분율 값을 0에서 1 사이의 소수로 변환
            deviation_sum += float(cell.value) * 0.01
        except (ValueError, TypeError):
            pass
rounded_deviation_sum = round(deviation_sum, 3)

# time 행의 가장 마지막 데이터와 첫번째 데이터를 빼서 spend_time에 저장
last_time = datetime.strptime(sheet.cell(row=sheet.max_row, column=2).value, "%H:%M:%S.%f")
first_time = datetime.strptime(sheet.cell(row=2, column=2).value, "%H:%M:%S.%f")
spend_time = last_time - first_time
total_sec = spend_time.total_seconds()

# 엑셀 파일의 열 수 계산 (제목 열을 제외)
total_columns = sheet.max_column - 1

# deviation_sum과 spend_time 출력
# print(f"Total Deviation : {rounded_deviation_sum}")
# print(f"Total Time : {spend_time}, (약 {total_sec} sec.)")

# deviation_mean 계산
deviation_mean = rounded_deviation_sum / total_columns
rounded_deviation_mean = round(deviation_mean * 100, 3)  # 백분율로 변환

print(f"운전자의 평균 차선 이탈률은 : {rounded_deviation_sum} % 입니다.")
